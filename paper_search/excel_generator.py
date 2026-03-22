import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _parse_papers(result_text: str):
    """
    Parse the LLM output and extract structured paper info.
    Returns a list of dicts with keys:
      number, title, verified, authors, year, journal,
      scholar_url, summary, apa, doi
    """
    papers = []
    # Split on --- separators
    blocks = re.split(r"\n---+\n", result_text)

    for block in blocks:
        block = block.strip()
        if not block:
            continue

        # Try to find paper number and title line: **[N] Title** [✅/⚠️]
        title_match = re.search(r"\*\*\[(\d+)\]\s*(.+?)\*\*\s*(✅|⚠️)?", block)
        if not title_match:
            continue

        number = title_match.group(1)
        title = title_match.group(2).strip()
        verified = title_match.group(3) or ""

        # Author / year / journal line
        meta_match = re.search(r"-\s*(.+?)\s*/\s*(\d{4})\s*/\s*(.+)", block)
        authors = meta_match.group(1).strip() if meta_match else ""
        year = meta_match.group(2).strip() if meta_match else ""
        journal = meta_match.group(3).strip() if meta_match else ""

        # Google Scholar URL
        scholar_match = re.search(r"(https://scholar\.google\.com/\S+)", block)
        scholar_url = scholar_match.group(1).strip() if scholar_match else ""

        # Summary (between **요약** and **APA 인용**)
        summary_match = re.search(r"\*\*요약\*\*\s*\n([\s\S]+?)(?=\*\*APA 인용\*\*|\Z)", block)
        summary = summary_match.group(1).strip() if summary_match else ""

        # APA citation
        apa_match = re.search(r"\*\*APA 인용\*\*\s*\n([\s\S]+)", block)
        apa = apa_match.group(1).strip() if apa_match else ""

        # DOI from APA
        doi_match = re.search(r"(https://doi\.org/\S+)", apa)
        doi = doi_match.group(1).strip() if doi_match else ""

        papers.append({
            "number": number,
            "title": title,
            "verified": verified,
            "authors": authors,
            "year": year,
            "journal": journal,
            "scholar_url": scholar_url,
            "summary": summary,
            "apa": apa,
            "doi": doi,
        })

    return papers


def _header_style():
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill("solid", fgColor="2C5F8A"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
    }


def _apply_header(cell, text):
    s = _header_style()
    cell.value = text
    cell.font = s["font"]
    cell.fill = s["fill"]
    cell.alignment = s["alignment"]
    cell.border = s["border"]


def _cell_style(wrap=True):
    return Alignment(vertical="top", wrap_text=wrap)


def generate_excel(output_path: str, keyword: str, date_str: str, result_text: str):
    papers = _parse_papers(result_text)

    wb = openpyxl.Workbook()

    # ---- Sheet 1: 논문목록 ----
    ws1 = wb.active
    ws1.title = "논문목록"
    ws1.row_dimensions[1].height = 22

    headers1 = ["번호", "제목", "저자", "연도", "학술지", "검증여부", "DOI", "Scholar URL"]
    for col, h in enumerate(headers1, 1):
        _apply_header(ws1.cell(row=1, column=col), h)

    for i, p in enumerate(papers, 2):
        row_data = [
            p["number"], p["title"], p["authors"], p["year"],
            p["journal"], p["verified"], p["doi"], p["scholar_url"],
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=i, column=col, value=val)
            cell.alignment = _cell_style()
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EAF2FB")

    col_widths1 = [6, 50, 30, 8, 30, 8, 40, 60]
    for col, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ---- Sheet 2: 요약모음 ----
    ws2 = wb.create_sheet("요약모음")
    ws2.row_dimensions[1].height = 22
    headers2 = ["번호", "제목", "200단어 요약"]
    for col, h in enumerate(headers2, 1):
        _apply_header(ws2.cell(row=1, column=col), h)

    for i, p in enumerate(papers, 2):
        for col, val in enumerate([p["number"], p["title"], p["summary"]], 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.alignment = _cell_style()
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EAF2FB")
        ws2.row_dimensions[i].height = 80

    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 45
    ws2.column_dimensions["C"].width = 80

    # ---- Sheet 3: APA인용 ----
    ws3 = wb.create_sheet("APA인용")
    ws3.row_dimensions[1].height = 22
    headers3 = ["번호", "완전한 APA 7th 인용문"]
    for col, h in enumerate(headers3, 1):
        _apply_header(ws3.cell(row=1, column=col), h)

    for i, p in enumerate(papers, 2):
        for col, val in enumerate([p["number"], p["apa"]], 1):
            cell = ws3.cell(row=i, column=col, value=val)
            cell.alignment = _cell_style()
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EAF2FB")
        ws3.row_dimensions[i].height = 50

    ws3.column_dimensions["A"].width = 6
    ws3.column_dimensions["B"].width = 100

    # ---- Meta info in each sheet header area ----
    for ws in [ws1, ws2, ws3]:
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ws[2]))
        meta_cell = ws.cell(row=1, column=1)
        meta_cell.value = f"논문 검색 결과  |  키워드: {keyword}  |  검색일: {date_str[:4]}.{date_str[4:6]}.{date_str[6:]}"
        meta_cell.font = Font(bold=True, color="2C5F8A", size=11)
        meta_cell.alignment = Alignment(horizontal="center", vertical="center")
        meta_cell.fill = PatternFill("solid", fgColor="DCE8F5")
        ws.row_dimensions[1].height = 20

    wb.save(output_path)
